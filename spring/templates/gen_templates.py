import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy

BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_MAP = {
    "fisico": ["CPF", "Nome", "RG", "Email", "Data de Nascimento",
               "Logradouro", "Número", "CEP", "Bairro", "Telefone",
               "Estado", "Cidade", "Principal", "Complemento"],
    "endereco": ["Logradouro", "Número", "CEP", "Bairro", "Telefone",
                 "Estado", "Cidade", "Principal", "Complemento"],
}

SHEET_NAMES = {
    "fisico": "Import Clientes Fisicos",
    "endereco": "Import Enderecos",
}

# --- Good data ---
GOOD_FISICO = [
    ["529.982.247-25", "João Silva Santos", "123456789", "joao.silva@email.com", "1990-05-15",
     "Rua das Flores", "123", "01001-000", "Centro", "(11) 91234-5678", "SP", "São Paulo", "Sim", "Apto 42"],
    ["088.945.771-68", "Maria Souza Oliveira", "987654321", "maria@email.com", "1985-08-22",
     "Avenida Paulista", "1000", "01310-100", "Bela Vista", "(11) 3234-5678", "SP", "São Paulo", "Sim", ""],
    ["584.340.390-19", "Carlos Pereira Lima", "456789123", "carlos@empresa.com", "2000-01-10",
     "Rua XV de Novembro", "500", "80020-310", "Centro", "(41) 91234-4321", "PR", "Curitiba", "Sim", "Sala 5"],
    ["714.516.450-87", "Ana Costa Rodrigues", "789123456", "ana.costa@exemplo.com", "1995-12-03",
     "Praça da Sé", "0", "01001-001", "Sé", "(11) 99876-5432", "SP", "São Paulo", "Sim", ""],
    ["414.333.190-53", "Pedro Almeida Neto", "321654987", "pedro@email.com", "1988-07-18",
     "Rua da Praia", "45", "90010-000", "Centro Histórico", "(51) 91234-5678", "RS", "Porto Alegre", "Sim", ""],
]

GOOD_ENDERECO = [
    ["Rua das Flores", "123", "01001-000", "Centro", "(11) 91234-5678", "SP", "São Paulo", "Sim", "Apto 42"],
    ["Avenida Paulista", "1000", "01310-100", "Bela Vista", "(11) 3234-5678", "SP", "São Paulo", "Sim", ""],
    ["Rua XV de Novembro", "500", "80020-310", "Centro", "(41) 91234-4321", "PR", "Curitiba", "Sim", "Sala 5"],
    ["Praça da Sé", "0", "01001-001", "Sé", "(11) 99876-5432", "SP", "São Paulo", "Sim", ""],
    ["Rua da Praia", "45", "90010-000", "Centro Histórico", "(51) 91234-5678", "RS", "Porto Alegre", "Sim", ""],
]

# --- Faulty data (each row has a different validation error) ---
FAULTY_FISICO = [
    ["123.456.789-01", "José Santos", "111222333", "jose@email.com", "1992-03-20",
     "Rua A", "10", "12345-678", "Centro", "(11) 91234-5678", "SP", "São Paulo", "Sim", ""],
    # Error: CPF inválido (123.456.789-01 fails check digits)

    ["529.982.247-25", "", "444555666", "pedro@email.com", "1990-05-15",
     "Rua B", "20", "12345-678", "Centro", "(11) 91234-5678", "SP", "São Paulo", "Sim", ""],
    # Error: Nome é obrigatório

    ["088.945.771-68", "Ana Oliveira", "777888999", "ana@email.com", "2000-01-10",
     "Rua C", "30", "1234-567", "Centro", "(11) 91234-5678", "SP", "São Paulo", "Sim", ""],
    # Error: CEP inválido (1234-567 has only 4 digits before dash)

    ["584.340.390-19", "Carlos Lima", "123456789", "carlos@email.com", "1995-07-15",
     "Rua D", "40", "87654-321", "Centro", "(11) 1234-5678", "SP", "São Paulo", "Sim", ""],
    # Error: Telefone inválido ((11) 1234-5678 -> 10 digits starting with 1, but landline must start [2-8])

    ["714.516.450-87", "Marina Costa", "555666777", "marina@email.com", "2099-01-01",
     "Rua E", "50", "11111-111", "Centro", "(21) 99876-5432", "RJ", "Rio de Janeiro", "Sim", ""],
    # Error: Data de nascimento não pode ser no futuro (2099-01-01)
]

FAULTY_ENDERECO = [
    ["", "123", "01001-000", "Centro", "(11) 91234-5678", "SP", "São Paulo", "Sim", ""],
    # Error: Logradouro é obrigatório

    ["Rua das Flores", "", "01001-000", "Centro", "(11) 91234-5678", "SP", "São Paulo", "Sim", ""],
    # Error: Número é obrigatório (empty cell -> parse null -> @NotNull)

    ["Avenida Paulista", "1000", "1234-567", "Bela Vista", "(11) 3234-5678", "SP", "São Paulo", "Sim", ""],
    # Error: CEP inválido

    ["Rua XV", "500", "80020-310", "Centro", "(11) 1111-1111", "PR", "Curitiba", "Sim", ""],
    # Error: Telefone inválido (only 8 digits - 11111111)

    ["Praça da Sé", "100", "01001-000", "", "(11) 99876-5432", "SP", "São Paulo", "Sim", ""],
    # Error: Bairro é obrigatório
]


def write_sheet(ws, headers, rows):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)


for kind in ["fisico", "endereco"]:
    for variant, data in [("bom", GOOD_FISICO if kind == "fisico" else GOOD_ENDERECO),
                          ("ruim", FAULTY_FISICO if kind == "fisico" else FAULTY_ENDERECO)]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAMES[kind]
        write_sheet(ws, HEADER_MAP[kind], data)
        fname = f"template-clientes-fisicos-{variant}.xlsx" if kind == "fisico" else f"template-enderecos-{variant}.xlsx"
        wb.save(f"/home/vaugtro/Documentos/Workspace/unika_estagio/spring/templates/{fname}")
        print(f"Generated: {fname}")
